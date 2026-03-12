"""Export service for generating reports in various formats.
Exports match the exact visual layout shown in the web application.
"""
import pandas as pd
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A3, A4, letter
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    KeepTogether, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ─── Color constants matching the web app ─────────────────────────
HDR_BLUE = colors.HexColor("#4472C4")
HDR_BLUE_BORDER = colors.HexColor("#3a63a8")
SUB_HDR_BG = colors.HexColor("#D6E4F0")
CAT_GREEN = colors.HexColor("#548235")
CAT_GREEN_BORDER = colors.HexColor("#3d6128")
CELL_BORDER = colors.HexColor("#B4C6E7")
COUNT_COL_BG = colors.HexColor("#D6E4F0")
AVG_COL_BG = colors.HexColor("#E8F0FE")
TOTAL_BG = colors.HexColor("#E2EFDA")
ROW_ALT = colors.HexColor("#F2F7FC")
TITLE_BG = colors.HexColor("#1f2937")
GREEN_TEXT = colors.HexColor("#1a5f2a")
BLUE_TEXT = colors.HexColor("#4472C4")
DARK_TEXT = colors.HexColor("#1f2937")

# Daily sold table colors
SOLD_HDR = colors.HexColor("#1f2937")
SOLD_CAT = colors.HexColor("#1a5f2a")
SOLD_DEPRE = colors.HexColor("#c2410c")
SOLD_MODEL = colors.HexColor("#1e40af")


class ExportService:
    """Service for exporting data to CSV, Excel, and PDF"""

    # ═══════════════════════════════════════════════════════════════
    # DAILY SOLD TABLE EXPORTS (Chart 1)
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def _flatten_grouped_data(daily_data):
        """Flatten grouped daily data into flat rows for CSV/Excel/PDF export."""
        if not daily_data or 'groups' not in daily_data:
            return []
        rows = []
        report_date = daily_data.get('date', '')
        for group in daily_data['groups']:
            category = group['category']
            for model_data in group['models']:
                model = model_data['name_model']
                entries = model_data.get('entries', [])
                if not entries:
                    rows.append({
                        'date': report_date,
                        'category': category,
                        'name_model': model,
                        'no': '',
                        'year_registered': '-',
                        'depreciation': '-',
                        'dealer_name': '-',
                        'price': '-'
                    })
                else:
                    for i, entry in enumerate(entries, 1):
                        price = entry.get('price')
                        price_str = f"${price:,.0f}" if price else '-'
                        rows.append({
                            'date': report_date,
                            'category': category,
                            'name_model': model if i == 1 else '',
                            'no': i,
                            'year_registered': entry['year_registered'],
                            'depreciation': entry['depreciation'],
                            'dealer_name': entry['dealer_name'],
                            'price': price_str
                        })
        return rows

    @staticmethod
    def to_daily_dataframe(daily_data):
        """Convert grouped daily data to DataFrame for export."""
        rows = ExportService._flatten_grouped_data(daily_data)
        if not rows:
            return pd.DataFrame(columns=['date', 'category', 'name_model', 'no',
                                         'year_registered', 'depreciation', 'dealer_name', 'price'])
        return pd.DataFrame(rows)

    # ── Daily CSV ──────────────────────────────────────────────────
    @staticmethod
    def export_daily_table_csv(daily_data):
        """Export daily sold table to CSV matching app layout."""
        rows = ExportService._flatten_grouped_data(daily_data)
        if not rows:
            df = pd.DataFrame(columns=['No', 'Category', 'Name & Model',
                                       'Year Reg', 'Depreciation', 'Dealer Name'])
        else:
            df = pd.DataFrame(rows)
            df = df.rename(columns={
                'no': 'No', 'category': 'Category', 'name_model': 'Name & Model',
                'year_registered': 'Year Reg', 'depreciation': 'Depreciation',
                'dealer_name': 'Dealer Name', 'price': 'Price'
            })
            df = df.drop(columns=['date'], errors='ignore')
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0)
        return output

    # ── Daily Excel ────────────────────────────────────────────────
    @staticmethod
    def export_daily_table_excel(daily_data):
        """Export daily sold table to Excel matching app layout with styling."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        rows = ExportService._flatten_grouped_data(daily_data)
        report_date = daily_data.get('date', '') if daily_data else ''

        output = BytesIO()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        ws.title = "Daily Sold Report"

        # Styles
        thin = Side(style="thin", color="E5E7EB")
        thick = Side(style="thin", color="374151")
        border_normal = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_header = Border(left=thick, right=thick, top=thick, bottom=thick)

        hdr_fill = PatternFill("solid", fgColor="1F2937")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        cat_fill = PatternFill("solid", fgColor="1A5F2A")
        cat_font = Font(bold=True, color="FFFFFF", size=10)
        model_font = Font(bold=True, color="1E40AF", size=9)
        depre_font = Font(bold=True, color="C2410C", size=9)
        normal_font = Font(size=9)
        empty_font = Font(italic=True, color="9CA3AF", size=9)

        # Title row
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1,
                             value=f"SGCarMart Daily Sold Report - {report_date}")
        title_cell.font = Font(bold=True, size=14, color="1A5F2A")
        title_cell.alignment = Alignment(horizontal="center")

        # Header row
        headers = ['No', 'Name & Model', 'Year Reg', 'Depreciation', 'Dealer Name', 'Price']
        col_widths = [6, 28, 10, 14, 22, 12]
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_header
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Data rows
        row_num = 4
        current_cat = None
        for r in rows:
            cat = r.get('category', '')
            # Category separator row
            if cat != current_cat:
                current_cat = cat
                ws.merge_cells(start_row=row_num, start_column=1,
                               end_row=row_num, end_column=6)
                cell = ws.cell(row=row_num, column=1, value=cat)
                cell.fill = cat_fill
                cell.font = cat_font
                cell.alignment = Alignment(horizontal="left")
                for c in range(1, 7):
                    ws.cell(row=row_num, column=c).border = border_normal
                    ws.cell(row=row_num, column=c).fill = cat_fill
                row_num += 1

            no_val = r.get('no', '')
            model = r.get('name_model', '')
            year = r.get('year_registered', '-')
            depre = r.get('depreciation', '-')
            dealer = r.get('dealer_name', '-')
            price = r.get('price', '-')

            is_empty = (no_val == '' and depre == '-')
            values = [no_val, model, year, depre, dealer, price]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border_normal
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 2 else "center",
                    vertical="center"
                )
                if is_empty:
                    cell.font = empty_font
                elif col_idx == 2 and model:
                    cell.font = model_font
                elif col_idx == 4:
                    cell.font = depre_font
                else:
                    cell.font = normal_font

            # Alternate row colors
            if row_num % 2 == 0:
                alt_fill = PatternFill("solid", fgColor="F9FAFB")
                for c in range(1, 7):
                    ws.cell(row=row_num, column=c).fill = alt_fill

            row_num += 1

        wb.save(output)
        output.seek(0)
        return output

    # ── Daily PDF ──────────────────────────────────────────────────
    @staticmethod
    def export_daily_table_pdf(daily_data, title="SGCarMart Daily Sold Report"):
        """Export daily sold table to PDF matching app layout."""
        rows = ExportService._flatten_grouped_data(daily_data)
        report_date = daily_data.get('date', '') if daily_data else ''

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4,
                                leftMargin=20, rightMargin=20,
                                topMargin=25, bottomMargin=25)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                     fontSize=14, textColor=GREEN_TEXT,
                                     spaceAfter=4)
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"<font size=8 color='#6b7280'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
            styles['Normal']))
        elements.append(Spacer(1, 10))

        if not rows:
            elements.append(Paragraph("No sold data available for this date.", styles['Normal']))
            doc.build(elements)
            output.seek(0)
            return output

        # Build table data with category rows
        col_widths = [28, 145, 52, 75, 130, 60]
        headers = ['No', 'Name & Model', 'Year', 'Depreciation', 'Dealer Name', 'Price']
        table_data = [headers]
        category_rows = [0]  # header row
        model_rows = []
        empty_rows = []

        current_cat = None
        row_idx = 1
        for r in rows:
            cat = r.get('category', '')
            if cat != current_cat:
                current_cat = cat
                table_data.append([cat, '', '', '', '', ''])
                category_rows.append(row_idx)
                row_idx += 1

            no_val = r.get('no', '')
            model = r.get('name_model', '')
            year = r.get('year_registered', '-')
            depre = r.get('depreciation', '-')
            dealer = r.get('dealer_name', '-')
            price = r.get('price', '-')

            is_empty = (no_val == '' and depre == '-')
            table_data.append([no_val, model, year, depre, dealer, price])
            if is_empty:
                empty_rows.append(row_idx)
            elif model:
                model_rows.append(row_idx)
            row_idx += 1

        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), SOLD_HDR),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, CELL_BORDER),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]

        # Category rows
        for cr in category_rows:
            if cr == 0:
                continue
            style_cmds.append(('SPAN', (0, cr), (-1, cr)))
            style_cmds.append(('BACKGROUND', (0, cr), (-1, cr), SOLD_CAT))
            style_cmds.append(('TEXTCOLOR', (0, cr), (-1, cr), colors.white))
            style_cmds.append(('FONTNAME', (0, cr), (-1, cr), 'Helvetica-Bold'))
            style_cmds.append(('FONTSIZE', (0, cr), (-1, cr), 8))
            style_cmds.append(('ALIGN', (0, cr), (-1, cr), 'LEFT'))

        # Model name rows - blue bold
        for mr in model_rows:
            style_cmds.append(('TEXTCOLOR', (1, mr), (1, mr), SOLD_MODEL))
            style_cmds.append(('FONTNAME', (1, mr), (1, mr), 'Helvetica-Bold'))

        # Depreciation column - orange
        for i in range(1, len(table_data)):
            if i not in category_rows:
                style_cmds.append(('TEXTCOLOR', (3, i), (3, i), SOLD_DEPRE))
                style_cmds.append(('FONTNAME', (3, i), (3, i), 'Helvetica-Bold'))

        # Alternating row colors
        for i in range(1, len(table_data)):
            if i not in category_rows and i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F9FAFB")))

        # Empty rows - italic gray
        for er in empty_rows:
            style_cmds.append(('TEXTCOLOR', (0, er), (-1, er), colors.HexColor("#9CA3AF")))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        output.seek(0)
        return output

    # ═══════════════════════════════════════════════════════════════
    # DEPRECIATION / UNITS TABLE EXPORTS (Chart 2)
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def _get_year_range():
        """Get year columns matching the web app."""
        current_year = datetime.now().year
        years = list(range(current_year, 2015, -1))  # 2026..2016
        year_labels = [str(y) for y in years] + ["2015 & Older"]
        year_keys = [str(y) for y in years] + ["2015 & Older"]
        return years + ["2015 & Older"], year_labels, year_keys

    @staticmethod
    def _flatten_depreciation_data(dep_data, categories, date_str=""):
        """Flatten depreciation pivot data into rows for CSV/Excel/PDF."""
        years, year_labels, year_keys = ExportService._get_year_range()

        rows = []
        for category, models in categories.items():
            for vehicle in models:
                vehicle_data = dep_data.get(vehicle, {})
                row = {"Category": category, "Vehicle": vehicle}
                total_units = 0
                for label, key in zip(year_labels, year_keys):
                    yd = vehicle_data.get(key) or vehicle_data.get(int(key) if key.lstrip('-').isdigit() else key, {})
                    low = yd.get("lowest", 0) if yd else 0
                    avg = yd.get("average", 0) if yd else 0
                    cnt = yd.get("unit", 0) if yd else 0
                    total_units += cnt
                    row[f"{label} LOW"] = f"${low:,}" if low else "-"
                    row[f"{label} AVG"] = f"${avg:,}" if avg else "-"
                    row[f"{label} COUNT"] = cnt if cnt else "-"
                row["TOTAL UNITS"] = total_units
                rows.append(row)
        return rows

    # ── Depreciation CSV ───────────────────────────────────────────
    @staticmethod
    def export_depreciation_csv(dep_data, categories, date_str=""):
        """Export depreciation table to CSV matching app layout."""
        rows = ExportService._flatten_depreciation_data(dep_data, categories, date_str)
        if not rows:
            df = pd.DataFrame(columns=["Category", "Vehicle", "TOTAL UNITS"])
        else:
            df = pd.DataFrame(rows)
        output = BytesIO()
        df.to_csv(output, index=False, encoding="utf-8")
        output.seek(0)
        return output

    # ── Depreciation Excel ─────────────────────────────────────────
    @staticmethod
    def export_depreciation_excel(dep_data, categories, date_str=""):
        """Export depreciation table to Excel matching exact app chart appearance."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        years, year_labels, year_keys = ExportService._get_year_range()
        year_group_count = len(year_labels)

        output = BytesIO()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        ws.title = "Depreciation"

        # Styles
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7")
        )
        thick_left = Border(
            left=Side(style="medium", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7")
        )

        hdr_fill = PatternFill("solid", fgColor="4472C4")
        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        sub_hdr_fill = PatternFill("solid", fgColor="D6E4F0")
        sub_hdr_font = Font(bold=True, color="1F2937", size=8)
        cat_fill = PatternFill("solid", fgColor="548235")
        cat_font = Font(bold=True, color="FFFFFF", size=10)
        vehicle_font = Font(bold=True, color="1E3A5F", size=9)
        low_font = Font(bold=True, color="1A5F2A", size=8)
        avg_font = Font(bold=True, color="4472C4", size=8)
        count_font = Font(bold=True, color="1F2937", size=8)
        total_font = Font(bold=True, color="1A5F2A", size=10)
        empty_font = Font(color="D1D5DB", size=8)
        title_fill = PatternFill("solid", fgColor="1F2937")
        title_font = Font(bold=True, color="FFFFFF", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")

        count_col_fill = PatternFill("solid", fgColor="D6E4F0")
        avg_col_fill = PatternFill("solid", fgColor="E8F0FE")
        total_col_fill = PatternFill("solid", fgColor="E2EFDA")
        alt_fill = PatternFill("solid", fgColor="F2F7FC")

        # Column layout: A=Vehicle, B..=year groups (3 cols each), last=TOTAL UNITS
        total_cols = 1 + (year_group_count * 3) + 1
        last_col = total_cols

        # ── Row 1: Title row ──
        d = datetime.strptime(date_str, "%Y-%m-%d") if date_str else datetime.now()
        months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                  'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        formatted_date = f"{d.day} {months[d.month - 1]}"

        cell = ws.cell(row=1, column=1, value=f"DATE: {formatted_date}")
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = left_align
        cell.border = thin_border

        ws.merge_cells(start_row=1, start_column=2,
                       end_row=1, end_column=total_cols - 1)
        title_cell = ws.cell(row=1, column=2,
                             value="D E P R E C I A T I O N   /   U N I T S")
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = center_align
        title_cell.border = thin_border

        ws.cell(row=1, column=total_cols).border = thin_border

        # ── Row 2: Year headers ──
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        cell = ws.cell(row=2, column=1)
        cell.fill = hdr_fill
        cell.border = thin_border

        col = 2
        for label in year_labels:
            ws.merge_cells(start_row=2, start_column=col,
                           end_row=2, end_column=col + 2)
            cell = ws.cell(row=2, column=col, value=label)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center_align
            cell.border = thin_border
            for c in range(col, col + 3):
                ws.cell(row=2, column=c).fill = hdr_fill
                ws.cell(row=2, column=c).border = thin_border
            col += 3

        # TOTAL UNITS header spans 2 rows
        ws.merge_cells(start_row=2, start_column=last_col,
                       end_row=3, end_column=last_col)
        cell = ws.cell(row=2, column=last_col, value="TOTAL\nUNITS")
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center_align
        cell.border = thin_border

        # ── Row 3: Sub-headers LOW/AVG/COUNT ──
        col = 2
        for _ in year_labels:
            for sub_label in ['LOW', 'AVG', 'COUNT']:
                cell = ws.cell(row=3, column=col, value=sub_label)
                cell.font = sub_hdr_font
                cell.alignment = center_align
                cell.border = thin_border
                if sub_label == 'COUNT':
                    cell.fill = count_col_fill
                else:
                    cell.fill = sub_hdr_fill
                col += 1

        # ── Data rows ──
        row_num = 4
        for category, models in categories.items():
            # Category row
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=total_cols)
            cell = ws.cell(row=row_num, column=1, value=category)
            cell.fill = cat_fill
            cell.font = cat_font
            cell.alignment = left_align
            for c in range(1, total_cols + 1):
                ws.cell(row=row_num, column=c).fill = cat_fill
                ws.cell(row=row_num, column=c).border = thin_border
            row_num += 1

            for v_idx, vehicle in enumerate(models):
                vehicle_data = dep_data.get(vehicle, {})
                is_alt = v_idx % 2 == 1
                bg = alt_fill if is_alt else None

                # Vehicle name
                cell = ws.cell(row=row_num, column=1, value=vehicle)
                cell.font = vehicle_font
                cell.alignment = left_align
                cell.border = thin_border
                if bg:
                    cell.fill = bg

                total_units = 0
                col = 2
                for label, key in zip(year_labels, year_keys):
                    yd = vehicle_data.get(key) or \
                         vehicle_data.get(int(key) if key.lstrip('-').isdigit() else key, {})
                    low = yd.get("lowest", 0) if yd else 0
                    avg_val = yd.get("average", 0) if yd else 0
                    cnt = yd.get("unit", 0) if yd else 0
                    total_units += cnt

                    has_data = low or cnt

                    # LOW
                    c = ws.cell(row=row_num, column=col,
                                value=f"${low:,}" if low else "-")
                    c.font = low_font if has_data else empty_font
                    c.alignment = center_align
                    c.border = thick_left
                    if bg:
                        c.fill = bg

                    # AVG
                    c = ws.cell(row=row_num, column=col + 1,
                                value=f"${avg_val:,}" if avg_val else "-")
                    c.font = avg_font if has_data else empty_font
                    c.alignment = center_align
                    c.border = thin_border
                    c.fill = avg_col_fill if not bg else PatternFill(
                        "solid", fgColor="E0EAFA" if is_alt else "E8F0FE")

                    # COUNT
                    c = ws.cell(row=row_num, column=col + 2,
                                value=cnt if cnt else "-")
                    c.font = count_font if has_data else empty_font
                    c.alignment = center_align
                    c.border = thin_border
                    c.fill = count_col_fill if not bg else PatternFill(
                        "solid", fgColor="C8D8E8" if is_alt else "D6E4F0")

                    col += 3

                # TOTAL UNITS
                c = ws.cell(row=row_num, column=last_col, value=total_units)
                c.font = total_font
                c.alignment = center_align
                c.border = thin_border
                c.fill = total_col_fill

                row_num += 1

        # Column widths
        ws.column_dimensions['A'].width = 22
        for c in range(2, total_cols):
            ws.column_dimensions[get_column_letter(c)].width = 10
        ws.column_dimensions[get_column_letter(last_col)].width = 10

        # Freeze panes
        ws.freeze_panes = 'B4'

        wb.save(output)
        output.seek(0)
        return output

    # ── Depreciation PDF ───────────────────────────────────────────
    @staticmethod
    def export_depreciation_pdf(dep_data, categories, date_str="", title="Depreciation / Units"):
        """Export depreciation table to PDF matching exact app chart appearance."""
        years, year_labels, year_keys = ExportService._get_year_range()
        year_group_count = len(year_labels)

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A3),
                                leftMargin=12, rightMargin=12,
                                topMargin=18, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()

        if not dep_data or all(not dep_data.get(v, {}) for cat_models in categories.values() for v in cat_models):
            elements.append(Paragraph(f"{title} - {date_str}", styles["Title"]))
            elements.append(Paragraph("No data available", styles["Normal"]))
            doc.build(elements)
            output.seek(0)
            return output

        # Parse date
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
        except (ValueError, TypeError):
            d = datetime.now()
        month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                       'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        formatted_date = f"{d.day} {month_names[d.month - 1]}"

        # Column widths: Vehicle + (LOW, AVG, COUNT) * yearGroups + TOTAL
        vehicle_w = 78
        low_w = 26
        avg_w = 26
        count_w = 20
        total_w = 28
        col_widths = [vehicle_w]
        for _ in range(year_group_count):
            col_widths.extend([low_w, avg_w, count_w])
        col_widths.append(total_w)

        total_cols = len(col_widths)

        # Build table data
        table_data = []

        # Row 0: Title row
        title_row = [f"DATE: {formatted_date}"]
        title_row.extend([''] * (total_cols - 2))
        title_row.append('')
        table_data.append(title_row)

        # Row 1: Year headers
        year_row = ['']  # vehicle col (will span to row 2)
        for label in year_labels:
            year_row.extend([label, '', ''])
        year_row.append('TOTAL\nUNITS')
        table_data.append(year_row)

        # Row 2: Sub-headers
        sub_row = ['']
        for _ in range(year_group_count):
            sub_row.extend(['LOW', 'AVG', 'CNT'])
        sub_row.append('')
        table_data.append(sub_row)

        # Track category rows and data rows
        category_row_indices = []
        data_row_vehicles = []  # (row_idx, v_idx_in_category)

        row_idx = 3
        for category, models in categories.items():
            cat_row = [category] + [''] * (total_cols - 1)
            table_data.append(cat_row)
            category_row_indices.append(row_idx)
            row_idx += 1

            for v_idx, vehicle in enumerate(models):
                vehicle_data = dep_data.get(vehicle, {})
                data_row = [vehicle]
                total_units = 0

                for label, key in zip(year_labels, year_keys):
                    yd = vehicle_data.get(key) or \
                         vehicle_data.get(int(key) if key.lstrip('-').isdigit() else key, {})
                    low = yd.get("lowest", 0) if yd else 0
                    avg_val = yd.get("average", 0) if yd else 0
                    cnt = yd.get("unit", 0) if yd else 0
                    total_units += cnt

                    data_row.append(f"${low:,}" if low else "-")
                    data_row.append(f"${avg_val:,}" if avg_val else "-")
                    data_row.append(str(cnt) if cnt else "-")

                data_row.append(str(total_units))
                table_data.append(data_row)
                data_row_vehicles.append((row_idx, v_idx))
                row_idx += 1

        t = Table(table_data, colWidths=col_widths, repeatRows=3)

        style_cmds = [
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.4, CELL_BORDER),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),

            # Title row (row 0)
            ('BACKGROUND', (0, 0), (0, 0), TITLE_BG),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 7),
            ('SPAN', (1, 0), (-2, 0)),
            ('ALIGN', (1, 0), (-2, 0), 'CENTER'),
            ('FONTNAME', (1, 0), (-2, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (-2, 0), 9),

            # Year header row (row 1)
            ('SPAN', (0, 1), (0, 2)),  # vehicle col spans 2 rows
            ('BACKGROUND', (0, 1), (0, 2), HDR_BLUE),
            ('SPAN', (-1, 1), (-1, 2)),  # total col spans 2 rows
            ('BACKGROUND', (-1, 1), (-1, 2), HDR_BLUE),
            ('TEXTCOLOR', (-1, 1), (-1, 2), colors.white),
            ('FONTNAME', (-1, 1), (-1, 2), 'Helvetica-Bold'),
            ('FONTSIZE', (-1, 1), (-1, 2), 6),
            ('ALIGN', (-1, 1), (-1, 2), 'CENTER'),

            # Sub-header row (row 2)
            ('FONTSIZE', (0, 2), (-1, 2), 5),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('ALIGN', (0, 2), (-1, 2), 'CENTER'),

            # Default data
            ('FONTNAME', (0, 3), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 3), (-1, -1), 5.5),
            ('ALIGN', (1, 3), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 3), (0, -1), 'LEFT'),
        ]

        # Year header spans (3 cols each)
        col = 1
        for _ in year_labels:
            style_cmds.append(('SPAN', (col, 1), (col + 2, 1)))
            style_cmds.append(('BACKGROUND', (col, 1), (col + 2, 1), HDR_BLUE))
            style_cmds.append(('TEXTCOLOR', (col, 1), (col + 2, 1), colors.white))
            style_cmds.append(('FONTNAME', (col, 1), (col + 2, 1), 'Helvetica-Bold'))
            style_cmds.append(('FONTSIZE', (col, 1), (col + 2, 1), 7))
            style_cmds.append(('ALIGN', (col, 1), (col + 2, 1), 'CENTER'))
            col += 3

        # Sub-header row coloring
        col = 1
        for _ in range(year_group_count):
            style_cmds.append(('BACKGROUND', (col, 2), (col, 2), SUB_HDR_BG))
            style_cmds.append(('BACKGROUND', (col + 1, 2), (col + 1, 2), SUB_HDR_BG))
            style_cmds.append(('BACKGROUND', (col + 2, 2), (col + 2, 2), COUNT_COL_BG))
            col += 3

        # Category rows
        for cr in category_row_indices:
            style_cmds.append(('SPAN', (0, cr), (-1, cr)))
            style_cmds.append(('BACKGROUND', (0, cr), (-1, cr), CAT_GREEN))
            style_cmds.append(('TEXTCOLOR', (0, cr), (-1, cr), colors.white))
            style_cmds.append(('FONTNAME', (0, cr), (-1, cr), 'Helvetica-Bold'))
            style_cmds.append(('FONTSIZE', (0, cr), (-1, cr), 7))
            style_cmds.append(('ALIGN', (0, cr), (-1, cr), 'LEFT'))

        # Data rows: coloring and alternating
        for (dr, v_idx) in data_row_vehicles:
            is_alt = v_idx % 2 == 1
            if is_alt:
                style_cmds.append(('BACKGROUND', (0, dr), (0, dr), ROW_ALT))

            # Vehicle name bold
            style_cmds.append(('FONTNAME', (0, dr), (0, dr), 'Helvetica-Bold'))
            style_cmds.append(('FONTSIZE', (0, dr), (0, dr), 5.5))
            style_cmds.append(('TEXTCOLOR', (0, dr), (0, dr), colors.HexColor("#1e3a5f")))

            # Per-year column coloring
            col = 1
            for _ in range(year_group_count):
                # LOW column
                if is_alt:
                    style_cmds.append(('BACKGROUND', (col, dr), (col, dr), ROW_ALT))
                style_cmds.append(('TEXTCOLOR', (col, dr), (col, dr), GREEN_TEXT))
                style_cmds.append(('FONTNAME', (col, dr), (col, dr), 'Helvetica-Bold'))

                # AVG column
                avg_bg = colors.HexColor("#E0EAFA") if is_alt else AVG_COL_BG
                style_cmds.append(('BACKGROUND', (col + 1, dr), (col + 1, dr), avg_bg))
                style_cmds.append(('TEXTCOLOR', (col + 1, dr), (col + 1, dr), BLUE_TEXT))
                style_cmds.append(('FONTNAME', (col + 1, dr), (col + 1, dr), 'Helvetica-Bold'))

                # COUNT column
                cnt_bg = colors.HexColor("#C8D8E8") if is_alt else COUNT_COL_BG
                style_cmds.append(('BACKGROUND', (col + 2, dr), (col + 2, dr), cnt_bg))
                style_cmds.append(('TEXTCOLOR', (col + 2, dr), (col + 2, dr), DARK_TEXT))
                style_cmds.append(('FONTNAME', (col + 2, dr), (col + 2, dr), 'Helvetica-Bold'))

                col += 3

            # TOTAL UNITS column
            style_cmds.append(('BACKGROUND', (-1, dr), (-1, dr), TOTAL_BG))
            style_cmds.append(('TEXTCOLOR', (-1, dr), (-1, dr), GREEN_TEXT))
            style_cmds.append(('FONTNAME', (-1, dr), (-1, dr), 'Helvetica-Bold'))
            style_cmds.append(('FONTSIZE', (-1, dr), (-1, dr), 7))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        output.seek(0)
        return output

    # ═══════════════════════════════════════════════════════════════
    # LEGACY EXPORTS (for backward compatibility)
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def to_dataframe(data):
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        columns_order = ['date', 'make_model', 'registered_year', 'depreciation',
                         'dealer_name', 'price']
        existing_columns = [col for col in columns_order if col in df.columns]
        return df[existing_columns]

    @staticmethod
    def export_to_csv(data):
        df = ExportService.to_dataframe(data)
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0)
        return output

    @staticmethod
    def export_to_excel(data):
        df = ExportService.to_dataframe(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Daily Report', index=False)
            worksheet = writer.sheets['Daily Report']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).apply(len).max(), len(str(col)))
                worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2
        output.seek(0)
        return output

    @staticmethod
    def export_to_pdf(data, title="SGCarMart Daily Report"):
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']))
        elements.append(Spacer(1, 12))
        if data:
            df = ExportService.to_dataframe(data)
            table_data = [df.columns.tolist()] + df.values.tolist()
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No data available", styles['Normal']))
        doc.build(elements)
        output.seek(0)
        return output

    @staticmethod
    def generate_summary(data):
        if not data:
            return {
                'total_listings': 0, 'highest_price': 0, 'lowest_price': 0,
                'average_price': 0, 'unique_models': 0
            }
        df = pd.DataFrame(data)
        prices = df['price'].dropna()
        return {
            'total_listings': len(df),
            'highest_price': float(prices.max()) if len(prices) > 0 else 0,
            'lowest_price': float(prices.min()) if len(prices) > 0 else 0,
            'average_price': float(prices.mean()) if len(prices) > 0 else 0,
            'unique_models': df['make_model'].nunique() if 'make_model' in df.columns else 0
        }
